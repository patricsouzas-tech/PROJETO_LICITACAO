import hashlib
from dataclasses import dataclass, field
from pathlib import Path

from ...core.logging import get_logger

logger = get_logger(__name__)

_NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


@dataclass
class TrechoExtraido:
    tipo_localizador: str
    texto_bruto: str
    pagina: int | None = None
    planilha: str | None = None
    celula_inicio: str | None = None
    celula_fim: str | None = None
    paragrafo: int | None = None
    tabela: str | None = None
    linha_tabela: int | None = None


@dataclass
class ExtracaoResultado:
    trechos: list[TrechoExtraido] = field(default_factory=list)
    precisa_ocr: bool = False
    erro: str | None = None


def _limpar(texto: str) -> str:
    return texto.replace("\u00a0", " ").strip()


def _sha256_texto(texto: str) -> str:
    return hashlib.sha256(texto.encode("utf-8")).hexdigest()


def extrair_pdf(caminho: Path) -> ExtracaoResultado:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ExtracaoResultado(erro="pypdf nao instalado")

    resultado = ExtracaoResultado()
    try:
        reader = PdfReader(str(caminho))
        for pagina_idx, page in enumerate(reader.pages, start=1):
            texto = _limpar(page.extract_text() or "")
            if not texto:
                resultado.precisa_ocr = True
                continue
            resultado.trechos.append(
                TrechoExtraido(
                    tipo_localizador="PAGINA", texto_bruto=texto, pagina=pagina_idx
                )
            )
        if not resultado.trechos:
            resultado.precisa_ocr = True
    except Exception as exc:  # noqa: BLE001
        resultado.erro = f"Falha ao extrair PDF: {exc}"
    return resultado


def extrair_docx(caminho: Path) -> ExtracaoResultado:
    try:
        from docx import Document
        from docx.oxml.ns import qn
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError:
        return ExtracaoResultado(erro="python-docx nao instalado")

    resultado = ExtracaoResultado()
    try:
        doc = Document(str(caminho))
        corpo = doc.element.body
        ordem = 0
        idx_tabela = 0
        for child in corpo.iterchildren():
            if child.tag == qn("w:p"):
                par = Paragraph(child, doc)
                texto = _limpar(par.text)
                if not texto:
                    continue
                ordem += 1
                resultado.trechos.append(
                    TrechoExtraido(
                        tipo_localizador="PARAGRAFO",
                        texto_bruto=texto,
                        paragrafo=ordem,
                    )
                )
            elif child.tag == qn("w:tbl"):
                idx_tabela += 1
                tabela = Table(child, doc)
                for linha_idx, row in enumerate(tabela.rows, start=1):
                    celulas = [_limpar(c.text) for c in row.cells]
                    texto = " | ".join(c for c in celulas if c)
                    if not texto:
                        continue
                    ordem += 1
                    resultado.trechos.append(
                        TrechoExtraido(
                            tipo_localizador="TABELA",
                            texto_bruto=texto,
                            tabela=str(idx_tabela),
                            linha_tabela=linha_idx,
                        )
                    )
    except Exception as exc:  # noqa: BLE001
        resultado.erro = f"Falha ao extrair DOCX: {exc}"
    return resultado


def extrair_xlsx(caminho: Path) -> ExtracaoResultado:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ExtracaoResultado(erro="openpyxl nao instalado")

    resultado = ExtracaoResultado()
    try:
        wb = load_workbook(str(caminho), read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                celulas = [
                    (c.coordinate, _limpar(str(c.value)))
                    for c in row
                    if c.value is not None
                ]
                celulas = [(coord, v) for coord, v in celulas if v]
                if not celulas:
                    continue
                primeira = celulas[0][0]
                ultima = celulas[-1][0]
                texto = " | ".join(v for _, v in celulas)
                resultado.trechos.append(
                    TrechoExtraido(
                        tipo_localizador="CELULA",
                        texto_bruto=texto,
                        planilha=sheet,
                        celula_inicio=primeira,
                        celula_fim=ultima,
                    )
                )
        wb.close()
    except Exception as exc:  # noqa: BLE001
        resultado.erro = f"Falha ao extrair XLSX: {exc}"
    return resultado


def extrair_txt(caminho: Path) -> ExtracaoResultado:
    resultado = ExtracaoResultado()
    try:
        raw = caminho.read_bytes()
        texto = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                texto = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if texto is None:
            texto = raw.decode("utf-8", errors="replace")
        linhas = [ln for ln in texto.splitlines()]
        for i, linha in enumerate(linhas, start=1):
            limpa = _limpar(linha)
            if not limpa:
                continue
            resultado.trechos.append(
                TrechoExtraido(
                    tipo_localizador="PARAGRAFO",
                    texto_bruto=limpa,
                    paragrafo=i,
                )
            )
    except Exception as exc:  # noqa: BLE001
        resultado.erro = f"Falha ao extrair TXT: {exc}"
    return resultado


_EXTENSOES_VALIDAS = {"pdf", "docx", "xlsx", "txt"}


def extrair(caminho: Path, extensao: str) -> ExtracaoResultado:
    ext = extensao.lower().lstrip(".")
    if ext not in _EXTENSOES_VALIDAS:
        return ExtracaoResultado(erro=f"Formato nao suportado: {ext}")
    if ext == "pdf":
        return extrair_pdf(caminho)
    if ext == "docx":
        return extrair_docx(caminho)
    if ext == "xlsx":
        return extrair_xlsx(caminho)
    if ext == "txt":
        return extrair_txt(caminho)
    return ExtracaoResultado(erro=f"Formato nao suportado: {ext}")
