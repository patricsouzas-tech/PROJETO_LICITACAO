from io import BytesIO

import pytest
from docx import Document as DocxDocument
from openpyxl import Workbook
from pypdf import PdfWriter

from licitacao.services.extraction.validate import (
    FormatoInvalidoError,
    validar_conteudo,
)


def _pdf_bytes() -> bytes:
    w = PdfWriter()
    w.add_blank_page(width=200, height=200)
    buf = BytesIO()
    w.write(buf)
    return buf.getvalue()


def _docx_bytes() -> bytes:
    d = DocxDocument()
    d.add_paragraph("paragrafo de teste")
    buf = BytesIO()
    d.save(buf)
    return buf.getvalue()


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "a"
    ws["B2"] = "b"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _txt_bytes() -> bytes:
    return "texto simples sem acentos binarios".encode("utf-8")


def test_pdf_valido_pass():
    validar_conteudo("pdf", _pdf_bytes())


def test_docx_valido_pass():
    validar_conteudo("docx", _docx_bytes())


def test_xlsx_valido_pass():
    validar_conteudo("xlsx", _xlsx_bytes())


def test_txt_valido_pass():
    validar_conteudo("txt", _txt_bytes())


def test_texto_renomeado_pdf_erro():
    with pytest.raises(FormatoInvalidoError):
        validar_conteudo("pdf", _txt_bytes())


def test_pdf_renomeado_txt_erro():
    with pytest.raises(FormatoInvalidoError):
        validar_conteudo("txt", _pdf_bytes())


def test_docx_renomeado_xlsx_erro():
    with pytest.raises(FormatoInvalidoError):
        validar_conteudo("xlsx", _docx_bytes())


def test_xlsx_renomeado_docx_erro():
    with pytest.raises(FormatoInvalidoError):
        validar_conteudo("docx", _xlsx_bytes())


def test_binario_nul_renomeado_txt_erro():
    with pytest.raises(FormatoInvalidoError):
        validar_conteudo("txt", b"texto\x00\x01\x02binario")


def test_extensao_nao_suportada_erro():
    with pytest.raises(FormatoInvalidoError):
        validar_conteudo("exe", b"qualquer")
